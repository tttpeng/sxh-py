#!/usr/bin/env python3
# -*- coding: utf-8 -*-
'''
Required
- requests (必须)
- pillow (可选)
Info
- author : "xchaoinfo"
- email  : "xchaoinfo@qq.com"
- date   : "2016.2.4"
Update
- name   : "wangmengcn"
- email  : "eclipse_sv@163.com"
- date   : "2016.4.21"
'''
import requests
import openpyxl
from bs4 import BeautifulSoup
from time import time
import datetime
import re


try:
    import cookielib
except:
    import http.cookiejar as cookielib
import re
import time
import os.path

try:
    from PIL import Image
except:
    pass

# 构造 Request headers
agent = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/54.0.2840.71 Safari/537.36"
headers = {
    "Host": "www.shanxinhui.com",
    "User-Agent": agent,
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
    "Pragma": "no-cache",
    "Upgrade-Insecure-Requests":"1",
    "Connection":"keep-alive"

}



allOrders = []

# 使用登录cookie信息
session = requests.session()
session.cookies = cookielib.LWPCookieJar('weibo_cookies.txt')
try:
    session.cookies.load(ignore_discard=True)
    print('get cookies')
    print(session.cookies)
except:
    print("Cookie 未能加载")
    
       



class Order(object):
    def __init__(self, Id, number1, number2):
        self.Id = Id
        self.number1 = number1
        self.number2 = number2


# 获取验证码
def get_captcha():
    t = str(int(time.time() * 1000))
    captcha_url = 'http://www.shanxinhui.com/User/Public/verify/' + t
    r = session.get(captcha_url, headers=headers)
    with open('captcha.jpg', 'wb') as f:
        f.write(r.content)
        f.close()
    # 用pillow 的 Image 显示验证码
    # 如果没有安装 pillow 到源代码所在的目录去找到验证码然后手动输入
    try:
        im = Image.open('captcha.jpg')
        im.show()
        im.close()
    except:
        print(u'请到 %s 目录找到captcha.jpg 手动输入' % os.path.abspath('captcha.jpg'))
    captcha = input("please input the captcha\n>")
    return captcha


def isLogin():
    # 通过查看用户个人信息来判断是否已经登录        print(session.cookies)
    print('----->')
    print(session.cookies)
    print('----->')
    url = "https://www.baidu.com/"
    login_page = session.get(url, headers=headers, allow_redirects=True)
    print(login_page.status_code)
    print(login_page.text)

    url = "http://www.shanxinhui.com/user/manager/listuseroutgo/p/1.html"
    login_page = session.get(url, headers=headers, allow_redirects=False)
    print(login_page.status_code)
    print(login_page.text)


    text = '2016-11-09 10:09:25'
    t_obj = time.strptime(text, "%Y-%m-%d %H:%M:%S")
    ts = time.mktime(t_obj)

    print(t_obj)
    # y = datetime.strptime(text, '%Y-%m-%d ')
    # z = datetime.now()
    # time_tuple = t_obj.timetuple()
    print(ts)

    # html = open('index.html', encoding="utf-8").read()
    # exculeExcel(html=login_page.text)


def getOrdersArrayResult(year,month,day):
    
    a= "%d-%d-%d 00:00:00" % (year,month,day)
    startTime = time.mktime(time.strptime(a,'%Y-%m-%d %H:%M:%S'))
    endTime = startTime + 60 * 60 * 24 - 1

    i = 1
    
    
    url = 'http://www.shanxinhui.com/user/manager/listuseroutgo/p/1.html'
    request_result = session.get(url,headers=headers, allow_redirects=False)

    html = request_result.text;
    prefectHtml = html.replace('</td>\n</td>', '</td>')
    soup = BeautifulSoup(prefectHtml, "html.parser")
    # soup = BeautifulSoup(login_page.text, "html.parser")
    # print(soup.find("table"))
    
    rowsNumber = soup.find("span", class_="rows")
    
    print('----')    
    print(rowsNumber.string)
    print('----')
    non_decimal = re.compile(r'[^\d.]+')
    number = non_decimal.sub('', rowsNumber.string)
    
    pageNumber = int(int(number)/18) + 1
    print(int(pageNumber) + 1)
    n = 1
    while n <  pageNumber + 1:
        getHTML(n)
        n += 1
    global allOrders        
    
    gggArray = []
    for dict in allOrders:
        if dict[3] > startTime and dict[3] < endTime:
            gggArray.append(dict)
            print(dict[3])
    
    newArray = []
    for array in gggArray:
        order, newArray = orderById(array[0], newArray)
        order.number1 = str(int(array[1]) + int(order.number1))
        order.number2 = str(int(array[2]) + int(order.number2))
    exculeExcel(newArray)
#    print(gggArray)
        
    


def getHTML(page):
    url = 'http://www.shanxinhui.com/user/manager/listuseroutgo/p/%d.html' % page;
    request_result = session.get(url,headers=headers, allow_redirects=False)
    
    print('html result -------')
    print(request_result.text)
    print('html result -------end')
    
    
    html = request_result.text;
    prefectHtml = html.replace('</td>\n</td>', '</td>')
    soup = BeautifulSoup(prefectHtml, "html.parser")
    
    rows = soup.find("table").find_all("tr")
    dict = {'id': [1, 2]}
    array = [dict]

    table = soup.find("table")
    result = makelist(table)

    result.pop()
    result.pop(0)
    newReuslt = [];
    
    for tempArray in result:
        print(tempArray[5])
        newReuslt.append([])
        typeString = tempArray[1]
        timeString = tempArray[5]
        t_obj = time.strptime(timeString, "%Y-%m-%d %H:%M:%S")
        ts = time.mktime(t_obj)
        numberString = tempArray[2]
        accountString = tempArray[3]        
        newReuslt[-1].insert(0, accountString)
        if typeString == '善种子':
            newReuslt[-1].insert(1, numberString)
        else:
            newReuslt[-1].insert(1, '0')
        if typeString == '善心币':
            newReuslt[-1].insert(2, numberString)
        else:
            newReuslt[-1].insert(2, '0')
        newReuslt[-1].insert(3,ts)
    print(newReuslt)
    global allOrders
    allOrders = allOrders + newReuslt





def getTimeOClockOfToday(year,month,day):
    import time

    t = time.localtime(time.time())
    time1 = time.mktime(time.strptime(time.strftime('%Y-%m-%d 00:00:00', t), '%Y-%m-%d %H:%M:%S'))


    return int(time1)


def exculeExcel(newArray):
    wb = openpyxl.load_workbook(filename='sxh.xlsx')
    ws = wb.create_sheet(title='zhenshide')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)

    ws.cell(row=1, column=1).value = '11.09 善种子100个'
    ws.cell(row=2, column=1).value = '      善心币16个'

    titles = ['姓名', '账号', '善种子', '善心币', '金额', '领导收钱', '转账金额']
    for idx, string in enumerate(titles):
        ws.cell(row=3, column=idx + 1).value = string

    for idx, order in enumerate(newArray):
        ws.cell(row=idx + 4, column=0 + 2).value = order.Id
        ws.cell(row=idx + 4, column=1 + 2).value = int(order.number1)
        ws.cell(row=idx + 4, column=2 + 2).value = int(order.number2)
        ws.cell(row=idx + 4, column=3 + 2).value = int(order.number1) * 300 + int(order.number2) * 100

    wb.save(filename='sxh.xlsx')

    # for idx, row in enumerate(newReuslt):
    #     for idxx, val in enumerate(row):
    #         ws.cell(row=idx + 1, column=idxx + 1).value = val
    #



    order = Order('0', '0', '0')
    order.Id = 'dd'
    order.number1 = '1';
    order.number2 = '2';
    print(order.Id)

    orderArray = [];
    # print(rows[0])
    # rows.pop()
    # for idx, row in enumerate(rows):
    #     cells = row.find_all("td")
    #     for idxx, val in enumerate(cells):
    #         if idxx == 3 :
    #             for dict in array:
    #                 if dict['id'] == val.get_text():
    #
    #
    #         # if val.get_text() == '善种子':
    #         #     print("这是一个山种子")
    #         # print(val.get_text())
    #         ws.cell(row=idx + 1, column=idxx + 1).value = val.get_text()
    #
    # wb.save(filename='sxh.xlsx')
    #
    #


    # print(rn)

    print('---')
    # if login_page.status_code == 200:
    #     return True
    # else:
    #     return False


def orderById(Id, orderArray):
    for order in orderArray:
        if order.Id == Id:
            return order, orderArray
    order = Order(Id, '0', '0')
    orderArray.append(order)
    return order, orderArray


def makelist(table):
    result = []
    allrows = table.findAll('tr')
    for row in allrows:
        result.append([])
        allcols = row.findAll('td')
        for col in allcols:
            thestrings = [s for s in col.findAll(text=True)]
            thetext = ''.join(thestrings)
            result[-1].append(thetext)
    return result


def login(secret, account):
    # 通过输入的用户名判断是否是手机号
    print("手机号登录 \n")
    post_url = 'http://www.shanxinhui.com/User/Public/login'
    postdata = {
        'password': secret,
        'username': account,
        'verify': 3169,
    }

    # 需要输入验证码后才能登录成功
    postdata["verify"] = get_captcha()
    login_page = session.post(post_url, data=postdata, headers=headers)
    print(login_page.text)
    # login_code = eval(login_page.text)
    # print(login_code)
    print(type(session.cookies))
    print(session.cookies)
    # session.cookies.save()
    session.cookies.save(ignore_discard=True)




def writeExcel():
    wb = openpyxl.load_workbook(filename='sxh.xlsx')
    ws = wb.create_sheet(title='hahah')
    ws.cell(row=1, column=1).value = 1232132
    wb.save(filename='sxh.xlsx')


try:
    input = raw_input
except:
    pass

if __name__ == '__main__':


    
    
    print('哈哈')
#    print(getTimeOClockOfToday(2016,11,10))    
    # writeExcel()
#    login('ss9988', 'csy9988')
    isLogin()
    getOrdersArrayResult(2016,11,11)
#        print('您已经登录')
        # else:
        #     account = input('请输入你的用户名\n>  ')
        #     secret = input("请输入你的密码\n>  ")
        #     login(secret, account)
